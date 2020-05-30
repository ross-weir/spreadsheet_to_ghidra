import json
from openpyxl import load_workbook

from src.clean_name import clean_name
from src.deserialize_signature import deserialize_signature
from src.gather_structs import gather_structs_from_entries


def main():
    wb = load_workbook(filename="../D2Common.xlsx", data_only=True)
    sheet = wb.worksheets[0]
    parsed_entries = []
    for row in sheet.iter_rows(9, sheet.max_row, 1, 4):
        ordinal, _, name, signature = row
        if name.value and signature.value and "-" not in name.value and type(signature.value) is str:
            parsed_entries.append({
                "ordinal": str(int(ordinal.value)),
                "name": clean_name(name.value),
                **deserialize_signature(signature.value),
            })
    structs = gather_structs_from_entries(parsed_entries)
    with open("../D2Common.json", "w") as f:
        json.dump({"functions": parsed_entries, "structs": structs}, f, indent=4)


if __name__ == "__main__":
    main()
